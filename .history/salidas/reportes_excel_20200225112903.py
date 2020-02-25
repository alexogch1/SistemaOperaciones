from django.shortcuts import render

from django.views .generic.base import TemplateView
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border,Font,PatternFill,Side
from django.utils import timezone

from django.views import generic


from .models import TiempoMuertoEnc, TiempoMuertonDet

class ReporteTmXls(TemplateView):
    def get (self, request, *args, **kwargs):

        query = TiempoMuertonDet.objects.all()
        wb = Workbook()

        ws = wb.active
        ws.tittle='Tiempos Muertos'


        #Establer el nombre del archivo
        nombre_archivo = "Reporte Tiempos Muertosw.xlsx"
        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Company'

        ws.merge_cells('B1:F1')

        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Department'

        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Reporte de Tiempos Muertos'

        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20


        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='Fecha'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Planta'

        ws['C7'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C7'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C7'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C7'].font = Font(name='calibri', size=11, bold=True)
        ws['C7']='Línea'

        ws['C8'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C8'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C8'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C8'].font = Font(name='calibri', size=11, bold=True)
        ws['C8']='supervisor'

        ws['C9'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C9'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C9'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C9'].font = Font(name='calibri', size=11, bold=True)
        ws['C9']='Turno'

        ws['C10'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C10'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C10'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C10'].font = Font(name='calibri', size=11, bold=True)
        ws['C10']='Causa'

        controlador = 7
        for q in query:

            ws.cell(row=controlador,column=7).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=7).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=7).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=7).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=7).value=str(q.causa)

            ws.cell(row=controlador,column=8).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=8).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=8).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=8).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=8).value=q.cantidad

            ws.cell(row=controlador,column=8).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=8).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=8).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=8).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=8).value=q.obs

            id_enc= q.tiempo_muerto_id


            query2 = TiempoMuertoEnc.objects.filter(id=id_enc)

            for x in query2:

                ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=2).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
                ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=2).value=x.fecha_produccion
                
                ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=3).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
                ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=3).value=str(x.planta)

                ws.cell(row=controlador,column=4).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=4).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=4).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
                ws.cell(row=controlador,column=4).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=4).value=str(x.linea)

                ws.cell(row=controlador,column=5).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=5).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=5).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
                ws.cell(row=controlador,column=5).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=5).value=str(x.supervisor)

                ws.cell(row=controlador,column=6).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=6).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=6).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
                ws.cell(row=controlador,column=6).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=6).value=str(x.turno)

            controlador +=1

        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response



