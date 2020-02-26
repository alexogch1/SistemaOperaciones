from django.shortcuts import render

from django.views .generic.base import TemplateView
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border,Font,PatternFill,Side
import datetime

from django.views import generic


from .models import TiempoMuertoEnc, TiempoMuertonDet, ProduccionEnc, ProduccionDet
from tmuertos.models import CausaTM, CausaTM

class ReporteTmXls(TemplateView):
    def get (self, request, *args, **kwargs):

        today =  datetime.datetime.now()
        today = today.strftime('%Y-%m-%d')

        query = TiempoMuertonDet.objects.all()
        wb = Workbook()

        ws = wb.active
        ws.tittle='Tiempos Muertos'


        #Establer el nombre del archivo
        nombre_archivo = str(today) + "Reporte Tiempos Muertosw.xlsx"
        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Mar Bran S.A. de C.V.'

        ws.merge_cells('B1:F1')

        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Innovación, Mejora Continua y Six Sigma'

        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Reporte de Tiempos Muertos'

 
        ws['G3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['G3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['G3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['G3'].font = Font(name='calibri', size=12, bold=True)
        ws['G3']='FECHA'

        ws['H3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['H3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['H3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['H3'].font = Font(name='calibri', size=12, bold=True)
        ws['H3']=today

        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=30
        ws.column_dimensions['F'].width=20
        ws.column_dimensions['G'].width=60
        ws.column_dimensions['H'].width=60
        ws.column_dimensions['G'].width=20
        ws.column_dimensions['J'].width=60


        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='Fecha'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Planta'

        ws['D6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['D6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['D6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['D6'].font = Font(name='calibri', size=11, bold=True)
        ws['D6']='Línea'

        ws['E6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['E6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['E6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['E6'].font = Font(name='calibri', size=11, bold=True)
        ws['E6']='Supervisor'

        ws['F6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['F6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['F6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['F6'].font = Font(name='calibri', size=11, bold=True)
        ws['F6']='Turno'


        ws['G6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['G6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['G6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['G6'].font = Font(name='calibri', size=11, bold=True)
        ws['G6']='Categoría'
        
        ws['H6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['H6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['H6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['H6'].font = Font(name='calibri', size=11, bold=True)
        ws['H6']='Causa'
        
        ws['I6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['I6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['I6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['I6'].font = Font(name='calibri', size=11, bold=True)
        ws['I6']='Tiempo (min)'

        ws['J6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['J6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['J6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['J6'].font = Font(name='calibri', size=11, bold=True)
        ws['J6']='Obs'

        controlador = 7
        for q in query:

            causa = q.causa
            query3 = CausaTM.objects.filter(descripcion=causa).first()

            categoria=query3.categoriaTM
            ws.cell(row=controlador,column=7).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=7).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=7).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=7).value=str(categoria)

            ws.cell(row=controlador,column=8).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=8).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=8).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=8).value=str(q.causa)

            ws.cell(row=controlador,column=9).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=9).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=9).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=9).value=q.cantidad

            ws.cell(row=controlador,column=10).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=10).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=10).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=10).value=q.obs

            id_enc= q.tiempo_muerto_id

            query2 = TiempoMuertoEnc.objects.filter(id=id_enc)

            for x in query2:

                ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=2).value=x.fecha_produccion
                
                ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=3).value=str(x.planta)

                ws.cell(row=controlador,column=4).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=4).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=4).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=4).value=str(x.linea)

                ws.cell(row=controlador,column=5).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=5).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=5).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=5).value=str(x.supervisor)

                ws.cell(row=controlador,column=6).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=6).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=6).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=6).value=str(x.turno)

            controlador +=1

        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReporteProdEmbXls(TemplateView):
    def get (self, request, *args, **kwargs):

        today =  datetime.datetime.now()
        today = today.strftime('%Y-%m-%d')

        query = ProduccionDet.objects.all()
        wb = Workbook()

        ws = wb.active
        ws.tittle='Producción Embolsados'


        #Establer el nombre del archivo
        nombre_archivo = str(today)  + "Reporte Producción Embolslados.xlsx"
        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Mar Bran S.A. de C.V.'

        ws.merge_cells('B1:F1')
        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Innovación, Mejora Continua y Six Sigma'
        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Reporte de Producción Embolsados'

 
        ws['G3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['G3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['G3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['G3'].font = Font(name='calibri', size=12, bold=True)
        ws['G3']='FECHA'

        ws['H3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['H3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['H3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['H3'].font = Font(name='calibri', size=12, bold=True)
        ws['H3']=today

        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=30
        ws.column_dimensions['F'].width=20
        ws.column_dimensions['G'].width=60
        ws.column_dimensions['H'].width=60
        ws.column_dimensions['G'].width=20
        ws.column_dimensions['J'].width=60


        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='Fecha'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Planta'

        ws['D6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['D6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['D6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['D6'].font = Font(name='calibri', size=11, bold=True)
        ws['D6']='Línea'

        ws['E6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['E6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['E6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['E6'].font = Font(name='calibri', size=11, bold=True)
        ws['E6']='Supervisor'

        ws['F6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['F6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['F6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['F6'].font = Font(name='calibri', size=11, bold=True)
        ws['F6']='Turno'


        ws['G6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['G6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['G6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['G6'].font = Font(name='calibri', size=11, bold=True)
        ws['G6']='Plantilla'
        
        ws['H6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['H6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['H6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['H6'].font = Font(name='calibri', size=11, bold=True)
        ws['H6']='Proc./term.'
        
        ws['I6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['I6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['I6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['I6'].font = Font(name='calibri', size=11, bold=True)
        ws['I6']='Producto'

        ws['J6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['J6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['J6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['J6'].font = Font(name='calibri', size=11, bold=True)
        ws['J6']='Peso (Lbs)'

        ws['K6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['K6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['K6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['K6'].font = Font(name='calibri', size=11, bold=True)
        ws['K6']='Cantidad (cajas)'

        ws['L6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['L6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['L6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['L6'].font = Font(name='calibri', size=11, bold=True)
        ws['L6']='Resto (lbs)'

        ws['M6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['M6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['M6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['M6'].font = Font(name='calibri', size=11, bold=True)
        ws['M6']='Producción (lbs)'     

        ws['N6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['N6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['N6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['N6'].font = Font(name='calibri', size=11, bold=True)
        ws['N6']='Utilizado (lbs)'  
    
        ws['O6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['O6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['O6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['O6'].font = Font(name='calibri', size=11, bold=True)
        ws['O6']='merma (%)'  

        controlador = 7
        for q in query:

            ws.cell(row=controlador,column=8).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=8).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=8).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=8).value=str(q.tproducto)

            ws.cell(row=controlador,column=9).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=9).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=9).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=9).value=str(q.producto)

            ws.cell(row=controlador,column=10).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=10).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=10).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=10).value=q.peso

            ws.cell(row=controlador,column=11).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=11).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=11).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=11).value=q.cantidad


            ws.cell(row=controlador,column=12).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=12).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=12).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=12).value=q.resto

            produccion = (float(q.peso)*float(q.cantidad)) + float(q.resto)
            ws.cell(row=controlador,column=13).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=13).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=13).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=13).value=produccion

            produccion = (float(q.peso)*float(q.cantidad)) + float(q.resto)
            ws.cell(row=controlador,column=13).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=13).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=13).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=13).value=produccion

            ws.cell(row=controlador,column=14).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=14).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=14).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=14).value=q.total_utilizado

            merma = float(produccion)/float(q.total_utilizado)+100
            ws.cell(row=controlador,column=15).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=15).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=15).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=15).value=merma

            id_enc= q.produccion_id
            query2 = ProduccionEnc.objects.filter(id=id_enc)
            for x in query2:

                ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=2).value=x.fecha_produccion

                ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=3).value=str(x.planta)

                ws.cell(row=controlador,column=4).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=4).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=4).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=4).value=str(x.linea)

                ws.cell(row=controlador,column=5).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=5).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=5).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=5).value=str(x.supervisor)


                ws.cell(row=controlador,column=6).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=6).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=6).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=6).value=str(x.turno)

                ws.cell(row=controlador,column=7).alignment= Alignment(horizontal='center', vertical='center')
                ws.cell(row=controlador,column=7).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                ws.cell(row=controlador,column=7).font = Font(name='calibri', size=11, bold=True)
                ws.cell(row=controlador,column=7).value=str(x.plantilla)
            controlador +=1






        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
