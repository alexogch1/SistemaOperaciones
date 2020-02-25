from django.shortcuts import render

from django.views .generic.base import TemplateView
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border,Font,PatternFill,Side
from django.utils import timezone

from django.views import generic


from .models import TiempoMuertoEnc, TiempoMuertonDet

class TiempoMuertoCompletoXls(generic.TemplateView):
    context_object_name='obj'
    queryset1 = TiempoMuertoEnc.objects.all()

    def get_context_data(self, **kwargs):
        context = super(TiempoMuertoCompletoXls, self).get_context_data(**kwargs)
        context['detalles'] = TiempoMuertonDet.objects.all()
        context['encabezado'] = self.queryset
        return context

    def get (self, *args, **kwargs):
        
        
        
        
        query = TiempoMuertoEnc.objects.all()
        print(query)
        wb = Workbook()
        

        ws = wb.active
        ws.tittle='TiemposMuertos'

        
        #Establer el nombre del archivo
        nombre_archivo = "TM.xlsx"

        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Mar Bran S.A. de C.V.'

        ws.merge_cells('B1:F1')

        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Innovación, Mejora Continua y Six Sigma'

        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Tipo de Cambio'

        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20

        
        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='Fecha'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Tipo de Cambio'

        controlador = 7
        for q in query:
            ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=2).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=2).value=q.fecha

            ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=3).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=3).value=q.tipo_cambio

            #contador+=1
            controlador +=1

        #Definir el tipo de resupuesta a dar    
        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#@login_required(login_url="/login/")
#@permission_required("catalogos.change_ingred",login_url="/login/")
def tc_inactivar(request,id):
    tc = TipoCambio.objects.filter(pk=id).first()

    if request.method=="POST":
        if tc:
            tc.estado = not tc.estado
            tc.save()
            return HttpResponse("OK")
        return HttpResponse("FAIL")
    
    return HttpResponse("FAIL")


class ReporteTC(TemplateView):
    def get (self, request, *args, **kwargs):
        print(request.GET.get('campo'))
        fecha_sel = request.GET.get('campo')
        fecha_sel_parse = parse(fecha_sel)
        print('fecha ',fecha_sel_parse.date())
        query = TipoCambio.objects.filter(fecha=fecha_sel)
        wb = Workbook()
        

        ws = wb.active
        ws.tittle='TipoCambio'

        
        #Establer el nombre del archivo
        nombre_archivo = "ReporteTC.xlsx"
        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Mar Bran S.A. de C.V.'

        ws.merge_cells('B1:F1')

        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Innovación, Mejora Continua y Six Sigma'

        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Tipo de Cambio'

        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20

        
        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='Fecha'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Tipo de Cambio'

        controlador = 7
        for q in query:
            ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=2).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=2).value=q.fecha

            ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws.cell(row=controlador,column=3).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=3).value=q.tipo_cambio

            #contador+=1
            controlador +=1

        #Definir el tipo de resupuesta a dar    
        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#@login_required(login_url="/login/")
#@permission_required("catalogos.change_ingred",login_url="/login/")
def tc_inactivar(request,id):
    tc = TipoCambio.objects.filter(pk=id).first()

    if request.method=="POST":
        if tc:
            tc.estado = not tc.estado
            tc.save()
            return HttpResponse("OK")
        return HttpResponse("FAIL")
    
    return HttpResponse("FAIL")