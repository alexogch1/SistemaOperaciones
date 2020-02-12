from django.shortcuts import render

from django.views .generic.base import TemplateView
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border,Font,PatternFill,Side
from django.utils import timezone

from django.views import generic

#from dateutil.parser import parse

from .models import CategoriaTM, CausaTM


class ReporteCatTMXls(TemplateView):
    def get (self, request, *args, **kwargs):
        today = timezone.now().date()
        query = CategoriaTM.objects.all()
        wb = Workbook()
        
        ws = wb.active
        ws.tittle='CatTiempoMuerto'

        
        #Establer el nombre del archivo
        nombre_archivo = "Categorias Tiempo Muerto.xlsx"
        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Mar Bran S.A. de C.V.'

        ws.merge_cells('B1:F1')

        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Innovación, Mejora Continua y Six Sigma'

        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Catalogo de Categorías Tiempo Muerto'
        
        ws['I3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['I3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['I3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['I3'].font = Font(name='calibri', size=12, bold=True)
        ws['I3']='Fecha '

        ws['J3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['J3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        #ws['J3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['J3'].font = Font(name='calibri', size=12, bold=True)
        ws['J3']=today


        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=30
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20
        ws.column_dimensions['J'].width=15

        
        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='id'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Descripcion'

        controlador = 7
        for q in query:
            ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=2).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=2).value=q.id_categoriaTM

            ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=3).fill = PatternFill(start_color='', end_color='', fill_type='solid')
            ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=3).value=q.descripcion

            #contador+=1
            controlador +=1

        #Definir el tipo de resupuesta a dar    
        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReporteCausaTMXls(TemplateView):
    def get (self, request, *args, **kwargs):
        today = timezone.now().date()
        query = CausaTM.objects.all()
        wb = Workbook()
        
        ws = wb.active
        ws.tittle='CausaTiempoMuerto'

        
        #Establer el nombre del archivo
        nombre_archivo = "Causas Tiempo Muerto.xlsx"
        ws['B1'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B1'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B1'].font = Font(name='calibri', size=12, bold=True)
        ws['B1']='Mar Bran S.A. de C.V.'

        ws.merge_cells('B1:F1')

        ws['B2'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B2'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B2'].font = Font(name='calibri', size=12, bold=True)
        ws['B2']='Innovación, Mejora Continua y Six Sigma'

        ws.merge_cells('B2:F2')
        ws['B3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['B3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font = Font(name='calibri', size=12, bold=True)
        ws['B3']='Catalogo de Causas Tiempo Muerto'
        
        ws['I3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['I3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws['I3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['I3'].font = Font(name='calibri', size=12, bold=True)
        ws['I3']='Fecha '

        ws['J3'].alignment= Alignment(horizontal='left', vertical='center')
        ws['J3'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        #ws['J3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['J3'].font = Font(name='calibri', size=12, bold=True)
        ws['J3']=today


        ws.merge_cells('B3:F3')

        ws.row_dimensions[1].height=20
        ws.row_dimensions[2].height=20
        ws.row_dimensions[3].height=20

        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=40
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20
        ws.column_dimensions['F'].width=40
        ws.column_dimensions['J'].width=15

        
        ws['B6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['B6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['B6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B6'].font = Font(name='calibri', size=11, bold=True)
        ws['B6']='id'


        ws['C6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['C6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['C6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C6'].font = Font(name='calibri', size=11, bold=True)
        ws['C6']='Descripcion'

        ws['D6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['D6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['D6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['D6'].font = Font(name='calibri', size=11, bold=True)
        ws['D6']='Tipo TM'

        ws['E6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['E6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['E6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['E6'].font = Font(name='calibri', size=11, bold=True)
        ws['E6']='Tolerancia'

        ws['F6'].alignment= Alignment(horizontal='center', vertical='center')
        ws['F6'].border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws['F6'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['F6'].font = Font(name='calibri', size=11, bold=True)
        ws['F6']='Categoria TM'

        controlador = 7
        for q in query:
            ws.cell(row=controlador,column=2).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=2).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=2).fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws.cell(row=controlador,column=2).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=2).value=q.id_causaTM

            ws.cell(row=controlador,column=3).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=3).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=3).fill = PatternFill(start_color='', end_color='', fill_type='solid')
            ws.cell(row=controlador,column=3).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=3).value=q.descripcion

            ws.cell(row=controlador,column=4).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=4).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=4).fill = PatternFill(start_color='', end_color='', fill_type='solid')
            ws.cell(row=controlador,column=4).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=4).value=q.tipo

            ws.cell(row=controlador,column=5).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=5).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=5).fill = PatternFill(start_color='', end_color='', fill_type='solid')
            ws.cell(row=controlador,column=5).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=5).value=q.tolerancia

            ws.cell(row=controlador,column=6).alignment= Alignment(horizontal='center', vertical='center')
            ws.cell(row=controlador,column=6).border =Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            #ws.cell(row=controlador,column=6).fill = PatternFill(start_color='', end_color='', fill_type='solid')
            ws.cell(row=controlador,column=6).font = Font(name='calibri', size=11, bold=True)
            ws.cell(row=controlador,column=6).value=str(q.categoriaTM)

            controlador +=1

        #Definir el tipo de resupuesta a dar    
        response = HttpResponse(content_type='application/ms-excel')
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
